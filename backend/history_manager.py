#!/usr/bin/env python3
"""
Libyana NPM - History Manager
Manages the historical Excel workbook with all sheets.
"""

import os
import logging
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from backend.site_processor import SITE_SUMMARY_HEADER
from backend.site_detail_processor import SITE_DETAIL_HEADER

logger = logging.getLogger(__name__)

# Default Excel file path
HISTORICAL_EXCEL = "Historical_Network_Data.xlsx"


class HistoryManager:
    """
    Manages the historical Excel workbook with multiple sheets.
    Handles append/update of data for SiteSummary, UserKPIs, Traffic, etc.
    """

    def __init__(self, excel_path=HISTORICAL_EXCEL):
        self.excel_path = excel_path
        self._ensure_workbook()

    def _ensure_workbook(self):
        """Create the Excel file with all required sheets if it doesn't exist."""
        if not os.path.exists(self.excel_path):
            logger.info(f"Creating new Excel workbook: {self.excel_path}")
            os.makedirs(os.path.dirname(self.excel_path) or '.', exist_ok=True)

            # Create all sheets in one go
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                # ... (all sheet creation code)
                pass  # Writer is automatically closed when exiting the with block

            logger.info(f"✅ Created workbook with all sheets at {self.excel_path}")
        else:
            # Verify the file is valid
            try:
                # Try to read the file
                pd.ExcelFile(self.excel_path)
            except Exception as e:
                logger.warning(f"Existing Excel file is corrupted, recreating: {e}")
                if os.path.exists(self.excel_path):
                    try:
                        os.remove(self.excel_path)
                        logger.info(f"Deleted corrupted file: {self.excel_path}")
                    except PermissionError:
                        logger.error("File is locked - close Excel and try again")
                        raise
                self._ensure_workbook()  # Recursive call to create new

    def _get_sheet(self, sheet_name):
        """Get a sheet as a DataFrame. Returns empty DataFrame if sheet doesn't exist."""
        try:
            if not os.path.exists(self.excel_path):
                return pd.DataFrame()

            # Read with pandas and ensure it's closed
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='openpyxl')
            return df
        except Exception as e:
            logger.debug(f"Sheet {sheet_name} not found or could not be read: {e}")
            return pd.DataFrame()

        
    def _write_sheet(self, sheet_name, df):
        """Write a DataFrame to a sheet, preserving other sheets."""
        if df is None or df.empty:
            logger.warning(f"No data to write to {sheet_name}")
            return

        # Ensure directory exists
        os.makedirs(os.path.dirname(self.excel_path) or '.', exist_ok=True)

        max_retries = 3
        for attempt in range(max_retries):
            try:
                # Read all existing sheets
                all_sheets = {}
                if os.path.exists(self.excel_path):
                    try:
                        # Read with pandas directly - this is more reliable
                        xls = pd.ExcelFile(self.excel_path, engine='openpyxl')
                        for name in xls.sheet_names:
                            if name != sheet_name:
                                all_sheets[name] = pd.read_excel(xls, sheet_name=name)
                                logger.debug(f"Read existing sheet: {name}")
                        xls.close()
                    except Exception as e:
                        logger.warning(f"Could not read existing sheets: {e}")
                        # If file is corrupted, try to read with openpyxl directly
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook(self.excel_path)
                            for name in wb.sheetnames:
                                if name != sheet_name:
                                    all_sheets[name] = pd.read_excel(self.excel_path, sheet_name=name)
                            wb.close()
                        except Exception as e2:
                            logger.error(f"Could not read any sheets: {e2}")
                            # If file is corrupted, recreate it
                            if os.path.exists(self.excel_path):
                                try:
                                    os.remove(self.excel_path)
                                    logger.info(f"Deleted corrupted file: {self.excel_path}")
                                    self._ensure_workbook()
                                except:
                                    pass

                # Add or update the sheet
                all_sheets[sheet_name] = df
                logger.info(f"Writing sheet {sheet_name} with {len(df)} rows")

                # Write back all sheets - use a fresh ExcelWriter
                # IMPORTANT: Use a new instance and ensure it's properly closed
                writer = pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='w')
                try:
                    for name, data in all_sheets.items():
                        data.to_excel(writer, sheet_name=name, index=False)
                        logger.debug(f"Wrote sheet: {name}")
                    writer.close()
                except Exception as e:
                    writer.close()
                    raise e

                logger.info(f"✅ Successfully wrote sheet {sheet_name} to {self.excel_path}")
                return True

            except PermissionError as e:
                if attempt < max_retries - 1:
                    logger.warning(f"Permission error (attempt {attempt + 1}/{max_retries}): {e}")
                    time.sleep(2)
                else:
                    logger.error(f"Failed after {max_retries} attempts: {e}")
                    raise
            except Exception as e:
                logger.error(f"Failed to write sheet {sheet_name}: {e}")
                if attempt < max_retries - 1:
                    logger.info(f"Retrying... (attempt {attempt + 1}/{max_retries})")
                    time.sleep(1)
                else:
                    raise

    def _append_with_dup_check(self, sheet_name, df, key_cols):
        """
        Append data with duplicate checking based on key columns.
        Returns (new_count, skipped_count)
        """
        if df is None or df.empty:
            return 0, 0

        try:
            # Ensure workbook exists
            self._ensure_workbook()

            # Get existing data (returns empty DataFrame if sheet doesn't exist)
            existing_df = self._get_sheet(sheet_name)

            if existing_df.empty:
                # Write directly
                self._write_sheet(sheet_name, df)
                logger.info(f"Created new sheet {sheet_name} with {len(df)} rows")
                return len(df), 0

            # Find available key columns
            available_keys = [col for col in key_cols if col in existing_df.columns and col in df.columns]

            if not available_keys:
                combined_df = pd.concat([existing_df, df], ignore_index=True)
                self._write_sheet(sheet_name, combined_df)
                logger.info(f"Appended {len(df)} rows to {sheet_name} (no duplicate check)")
                return len(df), 0

            # Create a set of existing keys
            existing_keys = set()
            for _, row in existing_df.iterrows():
                key = tuple(str(row.get(col, '')) for col in available_keys)
                existing_keys.add(key)

            # Filter new rows
            new_rows = []
            dup_count = 0
            for _, row in df.iterrows():
                key = tuple(str(row.get(col, '')) for col in available_keys)
                if key not in existing_keys:
                    new_rows.append(row)
                else:
                    dup_count += 1

            if not new_rows:
                logger.info(f"{sheet_name}: All {len(df)} rows already exist, skipping")
                return 0, len(df)

            # Append new rows
            new_df = pd.DataFrame(new_rows)
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            self._write_sheet(sheet_name, combined_df)
            logger.info(f"{sheet_name}: Appended {len(new_rows)} new rows (skipped {dup_count} duplicates)")

            return len(new_rows), dup_count

        except Exception as e:
            logger.error(f"Failed to append to {sheet_name}: {e}")
            raise
    # ---------- Site Summary Methods ----------
    def update_site_row(self, row_dict):
        """
        Update or append a row to the SiteSummary sheet.
        row_dict should contain all columns in SITE_SUMMARY_HEADER.
        """
        if not row_dict:
            logger.warning("No data to update in SiteSummary")
            return

        day = row_dict.get('day')
        if not day:
            logger.warning("Row has no 'day' field, cannot update SiteSummary")
            return

        try:
            # Ensure workbook exists
            self._ensure_workbook()

            df = self._get_sheet('SiteSummary')

            if not df.empty and day in df['day'].values:
                # Update existing row
                idx = df[df['day'] == day].index[0]
                for col in SITE_SUMMARY_HEADER:
                    df.at[idx, col] = row_dict.get(col, 0)
                logger.info(f"Updated SiteSummary for day {day}")
            else:
                # Append new row
                new_row = pd.DataFrame([row_dict])
                df = pd.concat([df, new_row], ignore_index=True)
                logger.info(f"Appended SiteSummary for day {day}")

            self._write_sheet('SiteSummary', df)
        except Exception as e:
            logger.error(f"Failed to update SiteSummary: {e}")
            raise

    def update_site_detail(self, df):
        """Replace the entire SiteDetail sheet with a new DataFrame."""
        if df is None or df.empty:
            logger.warning("No data to update in SiteDetail")
            return

        try:
            self._ensure_workbook()
            self._write_sheet('SiteDetail', df)
            logger.info(f"Updated SiteDetail with {len(df)} rows")
        except Exception as e:
            logger.error(f"Failed to update SiteDetail: {e}")
            raise

    # ---------- Traffic KPI Methods ----------
    def update_traffic_kpis(self, results_dict):
        """
        Update all traffic KPI sheets from the results dictionary.
        results_dict should have 'per_site' and 'whole_network' keys.
        """
        if not results_dict:
            logger.warning("No traffic KPI data to update")
            return

        self._ensure_workbook()

        total_new = 0
        total_skipped = 0

        # Update per-site traffic sheets
        if 'per_site' in results_dict:
            for sheet_name, df in results_dict['per_site'].items():
                if df is not None and not df.empty:
                    # Use Date and Site as key columns
                    key_cols = ['Date', 'Site']
                    new_count, skipped_count = self._append_with_dup_check(sheet_name, df, key_cols)
                    total_new += new_count
                    total_skipped += skipped_count
                    logger.info(f"Traffic per-site {sheet_name}: {new_count} new rows, {skipped_count} skipped")

        # Update whole network traffic sheets
        if 'whole_network' in results_dict:
            for sheet_name, df in results_dict['whole_network'].items():
                if df is not None and not df.empty:
                    # Use Date as key column
                    key_cols = ['Date']
                    # Create a combined sheet name for whole network
                    network_sheet_name = sheet_name.replace('Traffic_', 'Traffic_Network_')
                    new_count, skipped_count = self._append_with_dup_check(network_sheet_name, df, key_cols)
                    total_new += new_count
                    total_skipped += skipped_count
                    logger.info(
                        f"Traffic whole network {network_sheet_name}: {new_count} new rows, {skipped_count} skipped")

        logger.info(f"Traffic KPIs updated: {total_new} new rows, {total_skipped} duplicates skipped")
        return total_new, total_skipped

    # ---------- User KPIs Methods ----------
    def update_user_kpis(self, row_dict):
        """Update or append a row to UserKPIs sheet."""
        if not row_dict:
            return

        try:
            self._ensure_workbook()
            df = self._get_sheet('UserKPIs')
            day = row_dict.get('Day')

            if not day:
                logger.warning("User KPIs row missing Day")
                return

            if not df.empty and day in df['Day'].values:
                idx = df[df['Day'] == day].index[0]
                for col, value in row_dict.items():
                    if col in df.columns:
                        df.at[idx, col] = value
                logger.info(f"Updated UserKPIs for {day}")
            else:
                new_row = pd.DataFrame([row_dict])
                df = pd.concat([df, new_row], ignore_index=True)
                logger.info(f"Appended UserKPIs for {day}")

            self._write_sheet('UserKPIs', df)
        except Exception as e:
            logger.error(f"Failed to update UserKPIs: {e}")
            raise

    # ---------- Network KPI Methods ----------
    def update_network_kpis(self, results_dict):
        """
        Update all network KPI sheets from the results dictionary.
        Checks for duplicates before appending.
        """
        if not results_dict:
            logger.warning("No network KPI data to update")
            return

        self._ensure_workbook()
        total_new = 0
        total_skipped = 0

        for sheet_name, df in results_dict.items():
            if df is not None and not df.empty:
                key_cols = ['Date', 'Whole Network']
                new_count, skipped_count = self._append_with_dup_check(sheet_name, df, key_cols)
                total_new += new_count
                total_skipped += skipped_count
            else:
                logger.debug(f"No data to append to {sheet_name}")

        logger.info(f"Network KPIs updated: {total_new} new rows, {total_skipped} duplicates skipped")
        return total_new, total_skipped

    def update_network_kpi_sheet(self, sheet_name, df):
        """Update a specific network KPI sheet."""
        if df is None or df.empty:
            return
        self._ensure_workbook()
        key_cols = ['Date', 'Whole Network']
        self._append_with_dup_check(sheet_name, df, key_cols)

    # ---------- Cell KPI Methods ----------
    def update_cell_kpis(self, results_dict):
        """
        Update all cell KPI sheets from the results dictionary.
        Checks for duplicates before appending.
        """
        if not results_dict:
            logger.warning("No cell KPI data to update")
            return

        self._ensure_workbook()
        total_new = 0
        total_skipped = 0

        for sheet_name, df in results_dict.items():
            if df is not None and not df.empty:
                key_cols = ['Date', 'Cell Name']
                new_count, skipped_count = self._append_with_dup_check(sheet_name, df, key_cols)
                total_new += new_count
                total_skipped += skipped_count
                logger.info(f"Cell KPI {sheet_name}: {new_count} new rows, {skipped_count} skipped")
            else:
                logger.debug(f"No data to append to {sheet_name}")

        logger.info(f"Cell KPIs updated: {total_new} new rows, {total_skipped} duplicates skipped")
        return total_new, total_skipped

    # ---------- Packet Loss Methods ----------
    def update_packet_loss(self, df):
        """
        Update/append packet loss data.
        df should have columns: Date, GBSC, Adjacent Node Name, Adjacent Node Type,
        Adjacent Node ID, Hours_Above_5, Hours_100
        """
        if df is None or df.empty:
            return

        self._ensure_workbook()
        key_cols = ['Date', 'GBSC', 'Adjacent Node Name', 'Adjacent Node Type', 'Adjacent Node ID']
        self._append_with_dup_check('Packet_Loss', df, key_cols)
        logger.info(f"Updated Packet_Loss with {len(df)} rows")

    # ---------- EPT Methods ----------
    def update_ept_config(self, df):
        """Replace the entire EPT_Config sheet."""
        if df is None or df.empty:
            return

        try:
            self._ensure_workbook()
            self._write_sheet('EPT_Config', df)
            logger.info(f"Updated EPT_Config with {len(df)} rows")
        except Exception as e:
            logger.error(f"Failed to update EPT_Config: {e}")
            raise