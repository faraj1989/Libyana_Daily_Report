#!/usr/bin/env python3
"""
Libyana NPM - Excel Dashboard Generator
Generates a formatted Excel dashboard with all main KPIs for email reporting.
Default: Latest available day in Historical_Network_Data.xlsx
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

warnings.filterwarnings('ignore')


class ExcelDashboardGenerator:
    def __init__(self):
        self.excel_path = self._find_excel_file()
        self.data = {}
        self.available_dates = []
        self.load_data()

    def _find_excel_file(self):
        """Find Historical_Network_Data.xlsx"""
        if os.path.exists("Historical_Network_Data.xlsx"):
            return "Historical_Network_Data.xlsx"
        elif os.path.exists("output/Historical_Network_Data.xlsx"):
            return "output/Historical_Network_Data.xlsx"
        else:
            return None

    def load_data(self):
        """Load all sheets"""
        if not self.excel_path:
            print("❌ Historical_Network_Data.xlsx not found!")
            return

        try:
            xls = pd.ExcelFile(self.excel_path)
            for sheet in xls.sheet_names:
                try:
                    df = pd.read_excel(self.excel_path, sheet_name=sheet)
                    if not df.empty:
                        self.data[sheet] = df
                except:
                    pass

            # Get available dates from SiteSummary
            if 'SiteSummary' in self.data:
                df = self.data['SiteSummary']
                if 'day' in df.columns:
                    self.available_dates = sorted(df['day'].unique(), reverse=True)

            # Also from User_Summary
            if 'User_Summary' in self.data and not self.available_dates:
                df = self.data['User_Summary']
                if 'Date' in df.columns:
                    self.available_dates = sorted(df['Date'].unique(), reverse=True)

            print(f"✅ Loaded {len(self.data)} sheets")
            print(f"📅 Available dates: {len(self.available_dates)}")

        except Exception as e:
            print(f"❌ Error loading data: {e}")

    def get_data_for_date(self, date_str):
        """Extract metrics for a specific date"""
        metrics = {
            'date': date_str,
            '2G': {},
            '3G': {},
            '4G': {},
            'VoLTE': {},
            'users': {},
            'traffic': {},
            'summary': {}
        }

        # 1. Site Summary
        if 'SiteSummary' in self.data:
            df = self.data['SiteSummary']
            if 'day' in df.columns:
                row = df[df['day'] == date_str]
                if not row.empty:
                    row = row.iloc[0]
                    metrics['summary']['total_sites'] = row.get('Total Physical Sites (2G+3G+4G)', 0)
                    metrics['2G']['sites'] = row.get('2G physical sites', 0)
                    metrics['3G']['sites'] = row.get('3G physical sites', 0)
                    metrics['4G']['sites'] = row.get('4G physical sites', 0)

        # 2. User Summary
        if 'User_Summary' in self.data:
            df = self.data['User_Summary']
            if 'Date' in df.columns:
                row = df[df['Date'] == date_str]
                if not row.empty:
                    row = row.iloc[0]
                    metrics['users']['2g_cs'] = row.get('2G CS user', 0)
                    metrics['users']['3g_cs'] = row.get('3G CS user', 0)
                    metrics['users']['total_vlr'] = row.get('Total VLR Subscribers', 0)
                    metrics['users']['cs_roaming'] = row.get('Roaming CS (Almadar)', 0)
                    metrics['users']['roaming_2g_ps'] = row.get('Roaming 2G PS (Gb)', 0)
                    metrics['users']['roaming_3g_ps'] = row.get('Roaming 3G PS (Iu)', 0)
                    metrics['users']['roaming_4g_ps'] = row.get('Roaming 4G PS (S1)', 0)
                    metrics['users']['2g_ps'] = row.get('2G PS user', 0)
                    metrics['users']['3g_ps'] = row.get('3G PS user', 0)
                    metrics['users']['4g_ps'] = row.get('4G PS user', 0)
                    metrics['users']['volte'] = row.get('VoLTE user', 0)

        # 3. 4G Network
        if '4G_NWBH' in self.data:
            df = self.data['4G_NWBH']
            if 'Date' in df.columns:
                row = df[df['Date'] == date_str]
                if not row.empty:
                    row = row.iloc[0]
                    metrics['4G']['rrc_setup_success'] = row.get('RRC Setup Success Rate(%)', 0)
                    metrics['4G']['erab_setup_success'] = row.get('E-RAB Setup Success Rate', 0)
                    metrics['4G']['service_drop_rate'] = row.get('Service Drop Rate (All)', 0)
                    metrics['4G']['dl_throughput'] = row.get('User Downlink Average Throughput (Mbps)', 0)
                    metrics['4G']['ul_throughput'] = row.get('User Uplink Average Throughput (Mbps)', 0)
                    metrics['4G']['volte_cssr'] = row.get('VoLTE Setup Success Rate-ZM(%)', 0)
                    metrics['4G']['availability'] = row.get('Radio Network Availability Rate(%)', 0)
                    metrics['4G']['dl_prb_util'] = row.get('DL PRB Utilizing Rate(%)', 0)
                    metrics['4G']['ul_prb_util'] = row.get('UL PRB Utilizing Rate(%)', 0)
                    metrics['4G']['dl_traffic'] = row.get('Downlink Traffic Volume(GB)', 0)
                    metrics['4G']['ul_traffic'] = row.get('Uplink Traffic Volume (GB)', 0)
                    metrics['4G']['volte_traffic'] = row.get('VoLTE Traffic Volume (Erl)', 0)

        # 4. Traffic Network
        if 'Traffic_Network_4G' in self.data:
            df = self.data['Traffic_Network_4G']
            if 'Date' in df.columns:
                row = df[df['Date'] == date_str]
                if not row.empty:
                    row = row.iloc[0]
                    metrics['traffic']['4g_total'] = row.get('4G DL Traffic (GB)', 0)
                    metrics['traffic']['4g_volte'] = row.get('4G VoLTE Traffic (Erl)', 0)

        if 'Traffic_Network_2G' in self.data:
            df = self.data['Traffic_Network_2G']
            if 'Date' in df.columns:
                row = df[df['Date'] == date_str]
                if not row.empty:
                    row = row.iloc[0]
                    metrics['traffic']['2g_ps'] = row.get('2G PS Traffic (GB)', 0)
                    metrics['traffic']['2g_cs'] = row.get('2G CS Traffic (Erl)', 0)

        if 'Traffic_Network_3G' in self.data:
            df = self.data['Traffic_Network_3G']
            if 'Date' in df.columns:
                row = df[df['Date'] == date_str]
                if not row.empty:
                    row = row.iloc[0]
                    metrics['traffic']['3g_ps'] = row.get('3G PS Traffic (GB)', 0)
                    metrics['traffic']['3g_cs'] = row.get('3G CS Traffic (Erl)', 0)

        # Calculate totals
        metrics['users']['total_ps_users'] = (
                metrics['users'].get('2g_ps', 0) +
                metrics['users'].get('3g_ps', 0) +
                metrics['users'].get('4g_ps', 0)
        )

        metrics['users']['total_cs_users'] = (
                metrics['users'].get('2g_cs', 0) +
                metrics['users'].get('3g_cs', 0)
        )

        metrics['traffic']['total_ps_traffic'] = (
                metrics['traffic'].get('2g_ps', 0) +
                metrics['traffic'].get('3g_ps', 0) +
                metrics['traffic'].get('4g_total', 0)
        )

        metrics['traffic']['total_cs_traffic'] = (
                metrics['traffic'].get('2g_cs', 0) +
                metrics['traffic'].get('3g_cs', 0)
        )

        return metrics

    def get_latest_date(self):
        """Get latest available date"""
        if self.available_dates:
            return self.available_dates[0]
        return datetime.now().strftime('%Y-%m-%d')

    def validate_date(self, date_str):
        """Validate if date exists in data"""
        if date_str in self.available_dates:
            return True
        return False

    def generate_dashboard(self, date_str=None, output_file=None):
        """Generate formatted Excel dashboard"""
        if date_str is None:
            date_str = self.get_latest_date()

        # Validate date
        if not self.validate_date(date_str):
            print(f"⚠️ Warning: Date '{date_str}' not found in data!")
            print(f"📅 Available dates: {self.available_dates}")
            print(f"📌 Using latest date: {self.get_latest_date()}")
            date_str = self.get_latest_date()

        if output_file is None:
            output_file = f"Network_Dashboard_{date_str}.xlsx"

        print(f"📊 Generating dashboard for: {date_str}")
        print(f"📁 Output file: {output_file}")

        metrics = self.get_data_for_date(date_str)

        # Create DataFrame for summary
        summary_data = [
            ['📊 NETWORK PERFORMANCE SUMMARY', f'Date: {date_str}', 'EAST Region'],
            ['', '', ''],
            ['KPI Category', 'KPI Name', 'Value'],
            ['USER KPIs', 'Total PS Users', f"{metrics['users'].get('total_ps_users', 0):,}"],
            ['USER KPIs', 'Total CS Users', f"{metrics['users'].get('total_cs_users', 0):,}"],
            ['USER KPIs', '2G PS Users', f"{metrics['users'].get('2g_ps', 0):,}"],
            ['USER KPIs', '3G PS Users', f"{metrics['users'].get('3g_ps', 0):,}"],
            ['USER KPIs', '4G PS Users', f"{metrics['users'].get('4g_ps', 0):,}"],
            ['USER KPIs', 'VoLTE Users', f"{metrics['users'].get('volte', 0):,}"],
            ['USER KPIs', 'CS Roaming Users', f"{metrics['users'].get('cs_roaming', 0):,}"],
            ['USER KPIs', '2G PS Roaming', f"{metrics['users'].get('roaming_2g_ps', 0):,}"],
            ['USER KPIs', '3G PS Roaming', f"{metrics['users'].get('roaming_3g_ps', 0):,}"],
            ['USER KPIs', '4G PS Roaming', f"{metrics['users'].get('roaming_4g_ps', 0):,}"],
            ['', '', ''],
            ['TRAFFIC KPIs', 'Total PS Traffic (GB)', f"{metrics['traffic'].get('total_ps_traffic', 0):,.0f}"],
            ['TRAFFIC KPIs', 'Total CS Traffic (Erl)', f"{metrics['traffic'].get('total_cs_traffic', 0):,.0f}"],
            ['TRAFFIC KPIs', '2G PS Traffic (GB)', f"{metrics['traffic'].get('2g_ps', 0):,.2f}"],
            ['TRAFFIC KPIs', '3G PS Traffic (GB)', f"{metrics['traffic'].get('3g_ps', 0):,.2f}"],
            ['TRAFFIC KPIs', '4G PS Traffic (GB)', f"{metrics['traffic'].get('4g_total', 0):,.2f}"],
            ['TRAFFIC KPIs', '2G CS Traffic (Erl)', f"{metrics['traffic'].get('2g_cs', 0):,.2f}"],
            ['TRAFFIC KPIs', '3G CS Traffic (Erl)', f"{metrics['traffic'].get('3g_cs', 0):,.2f}"],
            ['TRAFFIC KPIs', '4G VoLTE Traffic (Erl)', f"{metrics['traffic'].get('4g_volte', 0):,.2f}"],
            ['', '', ''],
            ['4G KPIs', 'RRC Setup Success Rate (%)', f"{metrics['4G'].get('rrc_setup_success', 0):,.2f}"],
            ['4G KPIs', 'E-RAB Setup Success Rate (%)', f"{metrics['4G'].get('erab_setup_success', 0):,.2f}"],
            ['4G KPIs', 'Service Drop Rate (%)', f"{metrics['4G'].get('service_drop_rate', 0):,.3f}"],
            ['4G KPIs', 'Network Availability (%)', f"{metrics['4G'].get('availability', 0):,.2f}"],
            ['4G KPIs', 'DL Throughput (Mbps)', f"{metrics['4G'].get('dl_throughput', 0):,.3f}"],
            ['4G KPIs', 'UL Throughput (Mbps)', f"{metrics['4G'].get('ul_throughput', 0):,.3f}"],
            ['4G KPIs', 'VoLTE CSSR (%)', f"{metrics['4G'].get('volte_cssr', 0):,.2f}"],
            ['4G KPIs', 'DL PRB Utilization (%)', f"{metrics['4G'].get('dl_prb_util', 0):,.2f}"],
            ['4G KPIs', 'UL PRB Utilization (%)', f"{metrics['4G'].get('ul_prb_util', 0):,.2f}"],
            ['', '', ''],
            ['SITE KPIs', 'Total On-Air Sites', f"{metrics['summary'].get('total_sites', 0):,}"],
            ['SITE KPIs', '2G Sites', f"{metrics['2G'].get('sites', 0):,}"],
            ['SITE KPIs', '3G Sites', f"{metrics['3G'].get('sites', 0):,}"],
            ['SITE KPIs', '4G Sites', f"{metrics['4G'].get('sites', 0):,}"],
        ]

        df_summary = pd.DataFrame(summary_data[3:], columns=summary_data[2])

        # Create Excel with styling
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write summary data
            df_summary.to_excel(writer, sheet_name='Dashboard', index=False, startrow=3)

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Dashboard']

            # Add title
            worksheet.merge_cells('A1:C1')
            cell = worksheet['A1']
            cell.value = f"📊 Libyana Network Performance Dashboard - {date_str}"
            cell.font = Font(size=16, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.merge_cells('A2:C2')
            cell = worksheet['A2']
            cell.value = "EAST Region | Daily Report"
            cell.font = Font(size=12, italic=True, color="666666")
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Style headers
            for col in ['A', 'B', 'C']:
                cell = worksheet[f'{col}4']
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-fit columns - FIXED: handle merged cells properly
            for col_idx in ['A', 'B', 'C']:
                max_length = 0
                for row in worksheet.iter_rows(min_row=1, max_row=len(summary_data) + 2):
                    for cell in row:
                        if cell.column_letter == col_idx:
                            try:
                                if cell.value and not isinstance(cell, pd.core.generic.NDFrame):
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[col_idx].width = adjusted_width

            # Add borders (skip merged cells)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in worksheet.iter_rows(min_row=4, max_row=len(summary_data) + 2):
                for cell in row:
                    # Skip merged cells
                    if not isinstance(cell, pd.core.generic.NDFrame):
                        try:
                            cell.border = thin_border
                        except:
                            pass

            # Color-code values based on categories
            category_colors = {
                'USER KPIs': 'E8F4FD',
                'TRAFFIC KPIs': 'E8F8E8',
                '4G KPIs': 'FFF3E0',
                'SITE KPIs': 'F3E5F5'
            }

            for row in worksheet.iter_rows(min_row=5, max_row=len(summary_data) + 2):
                try:
                    category_cell = row[0]
                    if category_cell.value in category_colors:
                        color = category_colors[category_cell.value]
                        for cell in row:
                            if not isinstance(cell, pd.core.generic.NDFrame):
                                try:
                                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                except:
                                    pass
                except:
                    pass

            # Also create a simplified "copy-paste" sheet
            copy_sheet = workbook.create_sheet("Copy to Email")

            # Add plain text version for email
            copy_sheet['A1'] = f"Libyana Network Performance Dashboard - {date_str}"
            copy_sheet['A1'].font = Font(size=14, bold=True)

            row_num = 3
            for _, row in df_summary.iterrows():
                copy_sheet[f'A{row_num}'] = row['KPI Name']
                copy_sheet[f'B{row_num}'] = row['Value']
                row_num += 1

            # Auto-fit for copy sheet
            for col in ['A', 'B']:
                copy_sheet.column_dimensions[col].width = 40

        print(f"✅ Dashboard saved to: {output_file}")

        # Also generate plain text for easy copy-paste
        text_file = output_file.replace('.xlsx', '.txt')
        self._generate_text_summary(metrics, date_str, text_file)

        return output_file

    def _generate_text_summary(self, metrics, date_str, text_file):
        """Generate plain text summary for email"""
        lines = []
        lines.append("=" * 60)
        lines.append(f"LIBYANA NETWORK PERFORMANCE DASHBOARD - {date_str}")
        lines.append("=" * 60)
        lines.append("EAST Region | Daily Report")
        lines.append("")
        lines.append("-" * 60)
        lines.append("USER KPIs")
        lines.append("-" * 60)
        lines.append(f"Total PS Users        : {metrics['users'].get('total_ps_users', 0):,}")
        lines.append(f"Total CS Users        : {metrics['users'].get('total_cs_users', 0):,}")
        lines.append(f"2G PS Users           : {metrics['users'].get('2g_ps', 0):,}")
        lines.append(f"3G PS Users           : {metrics['users'].get('3g_ps', 0):,}")
        lines.append(f"4G PS Users           : {metrics['users'].get('4g_ps', 0):,}")
        lines.append(f"VoLTE Users           : {metrics['users'].get('volte', 0):,}")
        lines.append(f"CS Roaming Users      : {metrics['users'].get('cs_roaming', 0):,}")
        lines.append(f"2G PS Roaming         : {metrics['users'].get('roaming_2g_ps', 0):,}")
        lines.append(f"3G PS Roaming         : {metrics['users'].get('roaming_3g_ps', 0):,}")
        lines.append(f"4G PS Roaming         : {metrics['users'].get('roaming_4g_ps', 0):,}")
        lines.append("")
        lines.append("-" * 60)
        lines.append("TRAFFIC KPIs")
        lines.append("-" * 60)
        lines.append(f"Total PS Traffic (GB) : {metrics['traffic'].get('total_ps_traffic', 0):,.0f}")
        lines.append(f"Total CS Traffic (Erl): {metrics['traffic'].get('total_cs_traffic', 0):,.0f}")
        lines.append(f"2G PS Traffic (GB)    : {metrics['traffic'].get('2g_ps', 0):,.2f}")
        lines.append(f"3G PS Traffic (GB)    : {metrics['traffic'].get('3g_ps', 0):,.2f}")
        lines.append(f"4G PS Traffic (GB)    : {metrics['traffic'].get('4g_total', 0):,.2f}")
        lines.append(f"2G CS Traffic (Erl)   : {metrics['traffic'].get('2g_cs', 0):,.2f}")
        lines.append(f"3G CS Traffic (Erl)   : {metrics['traffic'].get('3g_cs', 0):,.2f}")
        lines.append(f"4G VoLTE Traffic (Erl): {metrics['traffic'].get('4g_volte', 0):,.2f}")
        lines.append("")
        lines.append("-" * 60)
        lines.append("4G KPIs")
        lines.append("-" * 60)
        lines.append(f"RRC Setup Success Rate  : {metrics['4G'].get('rrc_setup_success', 0):,.2f}%")
        lines.append(f"E-RAB Setup Success Rate: {metrics['4G'].get('erab_setup_success', 0):,.2f}%")
        lines.append(f"Service Drop Rate       : {metrics['4G'].get('service_drop_rate', 0):,.3f}%")
        lines.append(f"Network Availability    : {metrics['4G'].get('availability', 0):,.2f}%")
        lines.append(f"DL Throughput           : {metrics['4G'].get('dl_throughput', 0):,.3f} Mbps")
        lines.append(f"UL Throughput           : {metrics['4G'].get('ul_throughput', 0):,.3f} Mbps")
        lines.append(f"VoLTE CSSR              : {metrics['4G'].get('volte_cssr', 0):,.2f}%")
        lines.append(f"DL PRB Utilization      : {metrics['4G'].get('dl_prb_util', 0):,.2f}%")
        lines.append(f"UL PRB Utilization      : {metrics['4G'].get('ul_prb_util', 0):,.2f}%")
        lines.append("")
        lines.append("-" * 60)
        lines.append("SITE KPIs")
        lines.append("-" * 60)
        lines.append(f"Total On-Air Sites : {metrics['summary'].get('total_sites', 0):,}")
        lines.append(f"2G Sites           : {metrics['2G'].get('sites', 0):,}")
        lines.append(f"3G Sites           : {metrics['3G'].get('sites', 0):,}")
        lines.append(f"4G Sites           : {metrics['4G'].get('sites', 0):,}")
        lines.append("")
        lines.append("=" * 60)
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 60)

        with open(text_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        print(f"✅ Text summary saved to: {text_file}")


def main():
    """Main execution"""
    print("=" * 60)
    print("📊 Libyana NPM - Excel Dashboard Generator")
    print("=" * 60)

    generator = ExcelDashboardGenerator()

    if not generator.data:
        print("❌ No data loaded. Please ensure Historical_Network_Data.xlsx exists.")
        return

    # Ask user for date
    print("\n📅 Available dates:")
    for i, date in enumerate(generator.available_dates[:10]):
        print(f"  {i + 1}. {date}")

    if len(generator.available_dates) > 10:
        print(f"  ... and {len(generator.available_dates) - 10} more")

    # Default to latest
    latest = generator.get_latest_date()
    print(f"\n📌 Latest date: {latest}")

    choice = input(f"\nEnter date (YYYY-MM-DD) or press Enter for latest: ").strip()

    if not choice:
        date_str = latest
    else:
        date_str = choice

    # Validate date
    if not generator.validate_date(date_str):
        print(f"⚠️ Warning: Date '{date_str}' not found in data!")
        print(f"📅 Using latest date instead: {latest}")
        date_str = latest

    # Generate dashboard
    output_file = generator.generate_dashboard(date_str)

    print("\n" + "=" * 60)
    print("✅ Dashboard Generation Complete!")
    print(f"📁 Excel file: {output_file}")
    print(f"📁 Text file: {output_file.replace('.xlsx', '.txt')}")
    print("\n💡 To copy to email:")
    print("   1. Open the Excel file")
    print("   2. Go to 'Copy to Email' sheet")
    print("   3. Select all (Ctrl+A) and Copy (Ctrl+C)")
    print("   4. Paste into your email (Ctrl+V)")
    print("=" * 60)


if __name__ == "__main__":
    main()